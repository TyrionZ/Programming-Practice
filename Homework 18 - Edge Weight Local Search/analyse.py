import sys
import os
import re
import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active
rf = open("list", "r")
rows = 0

timePattern = re.compile(r"\d+\.\d+")
resultPattern = re.compile(r"result:(\d+)")

for line in rf:
	argv = line.split();
	if argv[0][0] == 'f':
		prefix = argv[0]
		ans = argv[1]
	else:
		maxSteps = argv[0]
		delta = argv[1]
		for i in range(5):
			rows += 1
			data = prefix + str(i + 1) + '.mis'

			resultFile = open("Results\\" + data + '-' + delta, "r")

			times = []
			results = []
			minResults = 1000000
			validTimes = []
			average = 0.0

			for l in resultFile:
			#	print(l)
				times += [float(timePattern.findall(l)[0])]
				results += [int(resultPattern.findall(l)[0])]
				if results[-1] < minResults:
					minResults = results[-1]

			for j in range(len(results)):
				if results[j] == minResults:
					validTimes.append(times[j])
					average += times[j]

			average /= len(validTimes)
			validTimes.sort()
			if len(validTimes) % 2 == 1:
				middle = validTimes[len(validTimes) >> 1]
			else:
				middle = (validTimes[len(validTimes) >> 1] + validTimes[(len(validTimes) >> 1) - 1]) / 2
			
			rate = str(len(validTimes)) + '/' + str(len(times))

			answer = str(minResults) + '/' + ans

			sheet.cell(row = rows, column = 1, value = data)
			sheet.cell(row = rows, column = 2, value = delta)
			sheet.cell(row = rows, column = 3, value = answer)
			sheet.cell(row = rows, column = 4, value = rate)
			sheet.cell(row = rows, column = 5, value = average)
			sheet.cell(row = rows, column = 6, value = middle)
			sheet.cell(row = rows, column = 7, value = str(2))


			resultFile.close()


data = "frb100-40.mis"
ans = "3900"
for i in range(5):
	delta = str(i + 1)
	rows += 1
	resultFile = open("Results\\" + data + '-' + delta, "r")

	times = []
	results = []
	minResults = 1000000
	validTimes = []
	average = 0.0

	for l in resultFile:
	#	print(l)
		times += [float(timePattern.findall(l)[0])]
		results += [int(resultPattern.findall(l)[0])]
		if results[-1] < minResults:
			minResults = results[-1]

	for j in range(len(results)):
		if results[j] == minResults:
			validTimes.append(times[j])
			average += times[j]

	average /= len(validTimes)
	validTimes.sort()
	if len(validTimes) % 2 == 1:
		middle = validTimes[len(validTimes) >> 1]
	else:
		middle = (validTimes[len(validTimes) >> 1] + validTimes[(len(validTimes) >> 1) - 1]) / 2
	
	rate = str(len(validTimes)) + '/' + str(len(times))

	answer = str(minResults) + '/' + ans

	sheet.cell(row = rows, column = 1, value = data)
	sheet.cell(row = rows, column = 2, value = delta)
	sheet.cell(row = rows, column = 3, value = answer)
	sheet.cell(row = rows, column = 4, value = rate)
	sheet.cell(row = rows, column = 5, value = average)
	sheet.cell(row = rows, column = 6, value = middle)
	sheet.cell(row = rows, column = 7, value = str(1))


	resultFile.close()

rf.close()
workbook.save("result.xlsx")